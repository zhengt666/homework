import numpy as np
from scipy.optimize import fsolve
from typing import List
import itertools
import networkx as nx
from domian import ProjectAnalysis
import matplotlib.pyplot as plt
import sys
import os,shutil
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from itertools import product

import logging
# 配置logging模块，设置日志级别、格式以及输出的文件名
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    filename='./out/output.log',
                    filemode='w')


def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def upload_excel_files():
    # 创建一个隐藏的主窗口
    root = tk.Tk()
    root.withdraw()
        
    # 设置上传目录
    upload_dir = os.path.join(os.getcwd(), 'projects')
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)
    else:
        shutil.rmtree(upload_dir)
        os.makedirs(upload_dir)
    # 弹出文件选择对话框，允许用户选择多个文件
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    
    # 如果没有选择文件，则退出函数
    if not file_path:
        exit()
    
    fixed_file_name = "uploaded_file.xlsx"
    # 构造完整的保存路径
    save_path = os.path.join(upload_dir, fixed_file_name)
        
    # 复制文件到上传目录
    try:
        with open(file_path, 'rb') as f_src, open(save_path, 'wb') as f_dst:
                f_dst.write(f_src.read())
            
        messagebox.showinfo("上传成功", f"文件 {os.path.basename(file_path)} 已成功上传到 {save_path}")
        
    except Exception as e:
        messagebox.showerror("上传失败", f"无法上传文件 {os.path.basename(file_path)}: {e}")


def get_os_front():
    if sys.platform == "win32":
        return "Microsoft YaHei"
    elif sys.platform == "darwin":
        return "PingFang SC"
    else:
        return "HeiTi"


# npv 可行性分析
def feasibility_check(project:ProjectAnalysis, discount_rate):
    logging.info(f"project {project.project_name} 可行性分析")
    npv_value = npv(project.project_cash_flows, discount_rate)
    logging.info(f"Net Present Value (NPV): {npv_value}")
    irr_value = irr(project.project_cash_flows)
    logging.info(f"Internal Rate of Return (IRR): {irr_value * 100:.2f}%")
    project.project_irr = irr_value
    project.project_npv = npv_value
    # 判断方案是否可行
    if npv_value > 0 and irr_value > discount_rate:
        logging.info(f"project {project.project_name} 方案可行。")
        return True
    else:
        logging.info(f"project {project.project_name} 方案不可行。")
        return False

# 计算净现值
def npv(cash_flows, discount_rate):
    n = len(cash_flows)
    for j in range(n):
        a = round(cash_flows[j] / (1 + discount_rate)**j,2)
    return round(sum(round(cash_flows[i] / (1 + discount_rate)**i,2) for i in range(n)),2)


# 计算单个项目的内部收益率，处理不同周期
def irr(cash_flows):
    def f(r):
        return sum(cash_flows[i] / (1 + r)**i for i in range(len(cash_flows)))
    return round(fsolve(f, 0.1)[0],4)

# 计算项目之间的差额内部收益率，扩展到不同周期
def diff_irr(cash_flows1, cash_flows2):
    max_period = max(len(cash_flows1), len(cash_flows2))
    extended_cash_flows1 = cash_flows1 + [0] * (max_period - len(cash_flows1))
    extended_cash_flows2 = cash_flows2 + [0] * (max_period - len(cash_flows2))
    combined_cash_flows = [a - b for a, b in zip(extended_cash_flows1, extended_cash_flows2)]
    return irr(combined_cash_flows)

# 计算优劣方案组合
def calc_rank(project_list: List[ProjectAnalysis], rate) -> List[ProjectAnalysis]:
    if len(project_list) <= 1:
        return project_list
    for project in project_list:
        calc_irr = irr(project.project_cash_flows)
        logging.info(f"Project {project.project_name} IRR: {calc_irr * 100:.2f}%")

    # 两两组合对比
    combinations = list(itertools.combinations(project_list, 2))
 
    diff_irrs = []
    # 打印结果
    for combo in combinations:
        logging.info(f"compare {combo[0].project_name} and {combo[1].project_name}")
        calc_diff_irr = diff_irr(combo[0].project_cash_flows, combo[1].project_cash_flows)
        logging.info(f"Difference IRR between Project {combo[0].project_name} and Project {combo[1].project_name}: {calc_diff_irr * 100:.2f}%")
        if calc_diff_irr < rate:
            diff_irrs.append((combo[1],combo[0]))
        else:
            diff_irrs.append((combo[0],combo[1]))
    
    # 创建有向图
    G = nx.DiGraph()
    # 添加边（对比关系）
    edges = diff_irrs
    G.add_edges_from(edges)
    
    try:
        # 尝试进行拓扑排序
        # 注意：在networkx的某些版本中，这个函数可能叫做lexicographical_topological_sort
        sorted_projects = list(nx.topological_sort(G))
        # 如果你的networkx版本没有topological_sort，尝试使用以下代码：
        # sorted_projects = list(nx.lexicographical_topological_sort(G))
        
        return sorted_projects
    except nx.NetworkXUnfeasible:
        # 如果图包含环，则拓扑排序不可行
        logging.info("The graph contains a cycle and cannot be topologically sorted.")

combinations_list:List[ProjectAnalysis] = []

# 计算组合的总收益
def calculate_return(project_list: List[ProjectAnalysis],total_investment):
    total_return = 0
    total_invest = 0
    chose_project:List[str] = []
    for project in project_list:
        jump = False
        for opposite in project.project_opposites:
            if opposite in chose_project:
                jump = True
                break
        if jump:
            continue
        total_invest += project.project_invest
        if total_invest >= total_investment:
            total_invest -= project.project_invest
            continue
        combinations_list.append(project)
        total_return += project.project_npv
        chose_project.append(project.project_name)
    return total_return

# 最优方案选择
def calc_best_combination(project_list: List[ProjectAnalysis],total_investment:float) -> List[ProjectAnalysis]:
    logging.info(f"Total return of best combination: {calculate_return(project_list,total_investment)}")
    logging.info(f"Best combination: { ','.join([obj.project_name for obj in combinations_list])}")
    return combinations_list

# 计算盈亏平衡
def calculate_break_even(project:ProjectAnalysis):
    fixed_cost, total_cost, production_capacity, selling_price = project.project_fixed_cost,project.project_total_cost,project.project_production_capacticy,project.project_selling_price
    # 计算变动成本
    variable_cost = (total_cost - fixed_cost) / production_capacity
    project.variable_cost = variable_cost
    
    # 计算盈亏平衡产量
    break_even_quantity = fixed_cost / (selling_price - variable_cost)
    
    # 计算盈亏平衡生产能力利率
    break_even_rate = (break_even_quantity / production_capacity) * 100
    
    # 计算盈亏平衡销售价格
    break_even_selling_price = variable_cost + (fixed_cost / production_capacity)
    
    # 计算盈亏平衡单位产品变动成本
    break_even_variable_cost = selling_price - (fixed_cost / production_capacity)
    
    # 计算经营安全率
    safety_margin = 1 - (break_even_quantity / production_capacity)

    # 打印结果
    logging.info(f"人流量: {round(break_even_quantity,0)} 人")
    logging.info(f"人均消费: {round(break_even_selling_price,2)} 元")
    logging.info(f"总固定成本: {round(fixed_cost,2)} 元")
    logging.info(f"单位产品变动成本: {round(variable_cost,2)} 元/件")
    
    
    logging.info(f"销售收入B=PQ= {round(selling_price*production_capacity,2)} 元")
    logging.info(f"总成本C=Cf+CvQ= {round(fixed_cost+variable_cost*production_capacity,2)} 元")
    logging.info(f"盈亏平衡点销量Q*=Cf/(P-Cv) {round(break_even_quantity,2)} 人")
    logging.info(f"盈亏平衡生产能力利用率: {round(break_even_rate,2)} %")
    logging.info(f"经营安全率: {safety_margin*100}%")
    
    if safety_margin > 0.5:
        logging.info ("经营安全")
    else:
        logging.info ("经营不安全")
    
    return {
        "break_even_quantity": break_even_quantity,
        "break_even_rate": break_even_rate,
        "break_even_selling_price": break_even_selling_price,
        "break_even_variable_cost": break_even_variable_cost,
        "safety_margin": safety_margin,
        "variable_cost": variable_cost
    } 

# 计算动态回收期
def dynamic_payback_period(project:ProjectAnalysis, discount_rate):
    """
    计算动态回收期。
    
    :param initial_investment: 初始投资额
    :param cash_flows: 现金流列表，按时间序列排列
    :param discount_rate: 折现率
    :return: 动态回收期
    """
    cumulative_npv = 0
    payback_period = 0
    for i, cash_flow in enumerate(project.project_cash_flows):
        cash_year = cash_flow / (1 + discount_rate) ** (i)
        cumulative_npv += cash_year
        if cumulative_npv >= 0:
            remaining_year = round(1 - (cumulative_npv / cash_year),2)
            payback_period = i - 1 + remaining_year
            break
    
    # 如果投资没有回收，则返回None
    if cumulative_npv < 0:
        return None
    logging.info(f"计算 project:{project.project_name} 的动态回收期为: {payback_period}")
    return payback_period

# 计算动态回收期
def dynamic_payback_period_by_cash_flows(cash_flows, discount_rate):
    """
    计算动态回收期。
    
    :param initial_investment: 初始投资额
    :param cash_flows: 现金流列表，按时间序列排列
    :param discount_rate: 折现率
    :return: 动态回收期
    """
    cumulative_npv = 0
    payback_period = 0
    for i, cash_flow in enumerate(cash_flows):
        cash_year = cash_flow / (1 + discount_rate) ** (i)
        cumulative_npv += cash_year
        if cumulative_npv >= 0:
            remaining_year = round(1 - (cumulative_npv / cash_year),2)
            payback_period = i - 1 + remaining_year
            break
    
    # 如果投资没有回收，则返回None
    if cumulative_npv < 0:
        return None
    return payback_period


# 计算静态回收期
def static_payback_period(project:ProjectAnalysis):
    """
    计算静态回收期。
    
    :param initial_investment: 初始投资额
    :param cash_flows: 现金流列表，按时间序列排列
    :param discount_rate: 折现率
    :return: 动态回收期
    """
    cumulative = 0
    payback_period = 0
    for i, cash_flow in enumerate(project.project_cash_flows):
        cash_year = cash_flow
        cumulative += cash_year
        if cumulative >= 0:
            remaining_year = round(1 - (cumulative / cash_year),2)
            payback_period = i - 1 + remaining_year
            break
    
    # 如果投资没有回收，则返回None
    if cumulative < 0:
        return None
    logging.info(f"计算 project:{project.project_name} 的静态回收期为: {payback_period}")
    return payback_period

# 绘制现金流量图
def draw_cash_flow(project:ProjectAnalysis):
    # 颜色设置
    arrow_color = 'black'
    axis_color = 'black'
    title_color = 'black'

    # 调整中文显示
    plt.rcParams['font.sans-serif'] = [get_os_front()]
    plt.rcParams['axes.unicode_minus'] = False
    # 设置绘图标题
    plt.title(f"Cash Flow Diagram of {project.project_name}", c=title_color)
    plt.ylabel("Funds (in ten thousand yuan)")

    # 设置初始变量值
    distance = project.distance
    logging.info(f"distance:{distance}")
    negative_array = [-x for x in project.project_cash_flows_out]
    cash_flow_array : List[List[int]] = [project.project_cash_flows_in,negative_array]

    # 根据列表绘制资金流量图
    for j in range(len(cash_flow_array)):
        for i in range(len(cash_flow_array[j])):
            if cash_flow_array[j][i] > 0:
                plt.arrow(i, 0, 0, cash_flow_array[j][i] - distance, fc=arrow_color, ec=arrow_color, shape="full", head_width=0.1, head_length=distance, overhang=0.5)
                plt.text(i + len(cash_flow_array) * 0.01, cash_flow_array[j][i] + distance, str(round(cash_flow_array[j][i], 2)))
            elif cash_flow_array[j][i] < 0:
                plt.arrow(i, 0, 0, cash_flow_array[j][i] + distance, fc=arrow_color, ec=arrow_color, shape="full", head_width=0.1, head_length=distance, overhang=0.5)
                plt.text(i + len(cash_flow_array) * 0.01, cash_flow_array[j][i] - distance, str(round(cash_flow_array[j][i], 2)))

    # 设置图表中各个元素的特征
    ax = plt.gca()
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['bottom'].set_position(('data', 0))
    plt.setp(ax.xaxis.get_majorticklabels(), ha="left")
    plt.arrow(-0.1, 0, len(cash_flow_array[0]) + 1.2, 0, fc=axis_color, ec=axis_color, shape="full", head_width=distance * 0.5, head_length=0.3, overhang=0.5)
    plt.yticks([])
    x_major_locator = plt.MultipleLocator(1)
    ax.xaxis.set_major_locator(x_major_locator)
    plt.xlim(-0.1, len(cash_flow_array[0]) + 1.4)
    plt.ylim(-15 * distance, 15 * distance)


    # 绘图
    plt.show()
    

# 定义净现值计算函数
def calculate_npv_single(K, B, C, r, n):
    return -K + (B - C) * (1 - (1 + r) ** -n) / r

# 定义单因素敏感性分析函数
def single_factor_sensitivity_analysis(factor, base_value, r, n, variation_range,K,B,C):
    npv_values = []
    for variation in variation_range:
        new_value = base_value * (1 + variation / 100)
        npv = calculate_npv_single(new_value, B, C, r, n) if factor == 'K' else calculate_npv_single(K , B if factor != 'B' else new_value, C if factor != 'C' else new_value, r, n)
        npv_values.append(npv)
    return npv_values

# 绘制单因素分析
def draw_single_factor(project:ProjectAnalysis,discount_rate):
    K,B,C,r,n = project.project_invest,project.project_cash_flows_in[1],project.project_cash_flows_out[1],discount_rate,project.project_period
    # 设置变动范围
    variation_range = np.arange(-20, 21, 5)

    # 分析投资额、年销售收入和年经营成本的敏感性
    npv_K = single_factor_sensitivity_analysis('K', K, r, n, variation_range,K,B,C)
    npv_B = single_factor_sensitivity_analysis('B', B, r, n, variation_range,K,B,C)
    npv_C = single_factor_sensitivity_analysis('C', C, r, n, variation_range,K,B,C)

    # 调整中文显示
    plt.rcParams['font.sans-serif'] = [get_os_front()]
    plt.rcParams['axes.unicode_minus'] = False
    # 绘制敏感性分析图
    plt.figure(figsize=(10, 6))
    plt.plot(variation_range, npv_K, label='Investment Sensitivity')
    plt.plot(variation_range, npv_B, label='Sales Revenue Sensitivity')
    plt.plot(variation_range, npv_C, label='Operating Cost Sensitivity')
    plt.xlabel('Variation (%)')
    plt.ylabel('NPV (10,000 yuan)')
    plt.title(f'Single Factor Sensitivity Analysis of {project.project_name}')
    plt.legend()
    plt.grid(True)
    plt.show()
    


# 定义净现值计算函数
def calculate_npv_multi(K, B, C, r, n, X, Y, Z):
    return -K * (1 + X) + (B * (1 + Y) - C * (1 + Z)) * (1 - (1 + r) ** -n) / r

# 绘制多因素分析
def draw_multi_factor(project:ProjectAnalysis,discount_rate):
    K,B,C,r,n = project.project_invest,project.project_cash_flows_in[1],project.project_cash_flows_out[1],discount_rate,project.project_period
    # 设置变动范围
    X_range = np.linspace(-0.2, 0.2, 100)  # 投资额变动范围 -20% 到 20%
    Y_range = np.linspace(-0.2, 0.2, 100)  # 年销售收入变动范围 -20% 到 20%
    Z_range = np.array([0.2, 0.1, 0, -0.1, -0.2])  # 经营成本变动范围 +20% 到 -20%

    # 创建网格
    X, Y = np.meshgrid(X_range, Y_range)

    # 计算NPV
    NPV = np.zeros_like(X)
    for i, z in enumerate(Z_range):
        NPV += calculate_npv_multi(K, B, C, r, n, X, Y, z) / len(Z_range)

    # 绘制等高线图
    # 调整中文显示
    plt.rcParams['font.sans-serif'] = [get_os_front()]
    plt.rcParams['axes.unicode_minus'] = False
    
    plt.figure(figsize=(10, 6))
    contour = plt.contour(X, Y, NPV, levels=[0], colors='black')
    plt.clabel(contour, inline=True, fontsize=8)
    plt.xlabel('Investment Change (%)')
    plt.ylabel('Sales Revenue Change (%)')
    plt.title(f'Multi-Factor Sensitivity Analysis of {project.project_name}')
    plt.grid(True)
    plt.show()
    
def get_base_data_openpyxl(file_path):
    """
    使用openpyxl库获取指定Excel文件第一个sheet的B1和B2单元格数据
    :param file_path: Excel文件的路径
    :return: B1和B2单元格的数据组成的元组，如果单元格为空则返回None
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        first_sheet = workbook.worksheets[0]

        b1_data = first_sheet['B1'].value
        b2_data = first_sheet['B2'].value * 10000

        return b1_data, b2_data
    except FileNotFoundError:
        logging.info(f"文件 {file_path} 未找到，请检查文件路径是否正确。")
        return None, None
    except Exception as e:
        logging.info(f"出现其他错误: {str(e)}")
        return None, None

def get_projects(file_path:str,discount_rate:float):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    all_sheets_except_first = workbook.worksheets[1:]
    project_list:List[ProjectAnalysis] = []
    project_opposites:List[List[ProjectAnalysis]] = []
    project_opposites_name = []
    for sheet in all_sheets_except_first:
        logging.info(f"读取sheet: {sheet.title}")
        project_name_1 = sheet['B1'].value
        project_proid_1 = sheet['B2'].value
        project_inflow_1 = []
        project_outflow_1 = []
        project_opposite = []
        for col in range(2, 3+project_proid_1):  # 对应第二列到第七列，在openpyxl中列索引从1开始
            in_cell = sheet.cell(row=5, column=col)
            project_inflow_1.append(in_cell.value)
            out_cell = sheet.cell(row=9, column=col)
            project_outflow_1.append(out_cell.value)
        project_fixed_cost_1 = sheet['B19'].value
        project_total_cost_1 = sheet['B18'].value
        project_production_capacticy_1 = sheet['B16'].value
        project_selling_price_1 = sheet['B17'].value
        project_opposites_1 = sheet['Q2'].value
        project1 = ProjectAnalysis(project_name_1,project_inflow_1,project_outflow_1,project_fixed_cost_1,project_total_cost_1,project_production_capacticy_1,project_selling_price_1,discount_rate,[project_opposites_1])
        
        project_name_2 = sheet['Q1'].value
        project_proid_2 = sheet['Q2'].value
        project_inflow_2 = []
        project_outflow_2 = []
        for col in range(17, 18+project_proid_2):  # 对应第二列到第七列，在openpyxl中列索引从2开始
            in_cell = sheet.cell(row=5, column=col)
            project_inflow_2.append(in_cell.value)
            out_cell = sheet.cell(row=9, column=col)
            project_outflow_2.append(out_cell.value)
        project_fixed_cost_2 = sheet['Q19'].value
        project_total_cost_2 = sheet['Q18'].value
        project_production_capacticy_2 = sheet['Q16'].value
        project_selling_price_2 = sheet['Q17'].value
        project_opposites_2 = sheet['B2'].value
        project2 = ProjectAnalysis(project_name_2,project_inflow_2,project_outflow_2,project_fixed_cost_2,project_total_cost_2,project_production_capacticy_2,project_selling_price_2,discount_rate,[project_opposites_2])
        
        project_list.append(project1)
        project_list.append(project2)
        project_opposite.append(project1)
        project_opposite.append(project2)
        project_opposites.append(project_opposite)
        project_opposites_name.append(sheet.title)
    return project_list,project_opposites,project_opposites_name


def dynamic_payback_period_difference(scheme1_cash_flows, scheme2_cash_flows, discount_rate):
    """
    计算两个方案的差额投资动态回收期

    参数:
    scheme1_cash_flows (list 或 np.array): 方案1各期的现金流量序列，例如[初始投资（第0期现金流量）, 第1期现金流量, 第2期现金流量,...]
    scheme2_cash_flows (list 或 np.array): 方案2各期的现金流量序列，例如[初始投资（第0期现金流量）, 第1期现金流量, 第2期现金流量,...]
    discount_rate (float): 折现率，例如0.1表示10%

    返回:
    float: 差额投资动态回收期（年数，如果回收期不是整数年会返回小数表示），若在给定现金流情况下无法收回投资则返回None
    """
    # 确保两个方案的现金流量期数一致
    if len(scheme1_cash_flows)!= len(scheme2_cash_flows):
        raise ValueError("两个方案的现金流量期数不一致，请检查输入数据。")

    net_cash_flows = np.array([scheme1_cash_flows[i] - scheme2_cash_flows[i] for i in range(len(scheme1_cash_flows))])
    return dynamic_payback_period_by_cash_flows(net_cash_flows,discount_rate)

def break_even_analysis_extended(fixed_costs, variable_cost_per_unit, selling_price_per_unit, max_units=1000):
    """
    扩展的盈亏平衡分析，包含不同产量下利润情况分析及可视化展示

    参数:
    fixed_costs (float): 固定成本，即不随产量变化的成本（例如设备租金、厂房租金等）
    variable_cost_per_unit (float): 单位可变成本，即每生产一单位产品需要额外花费的成本（例如原材料、人工等按单位计算的成本）
    selling_price_per_unit (float): 单位产品售价
    max_units (int): 分析的最大产量范围，默认值为1000

    返回:
    float: 盈亏平衡点的产量，即达到收支平衡时需要销售的产品数量，若售价小于等于单位可变成本则返回None表示无法实现盈利
    """
    if selling_price_per_unit <= variable_cost_per_unit:
        return None

    # 计算盈亏平衡点
    break_even_point = fixed_costs / (selling_price_per_unit - variable_cost_per_unit)

    # 生成不同产量下的成本、收入和利润数据
    units = range(0, max_units + 1)
    total_costs = [fixed_costs + variable_cost_per_unit * unit for unit in units]
    total_revenues = [selling_price_per_unit * unit for unit in units]
    profits = [revenue - cost for revenue, cost in zip(total_revenues, total_costs)]

    # 绘制成本、收入和利润曲线
    plt.plot(units, total_costs, label='总成本')
    plt.plot(units, total_revenues, label='总收入')
    plt.plot(units, profits, label='总利润')
    plt.axvline(x=break_even_point, color='r', linestyle='--', label='盈亏平衡点')
    plt.xlabel('产量（单位）')
    plt.ylabel('金额')
    plt.title('盈亏平衡分析')
    plt.legend()
    plt.show()

    return break_even_point

def generate_combinations(elements: List[ProjectAnalysis]):
    """
    使用itertools.product生成给定元素的所有组合（包含元素选或不选的情况）
    :param elements: 输入的元素列表
    :return: 生成的所有组合的列表，每个组合也是一个列表
    """
    all_combinations: List[List[ProjectAnalysis]] = []
    for combination in product([True, False], repeat=len(elements)):
        current_combination = [elem for elem, included in zip(elements, combination) if included]
        all_combinations.append(current_combination)
    return all_combinations