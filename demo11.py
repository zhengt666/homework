import os
from openpyxl import load_workbook

def get_excel_values(directory):
    values = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):  # 根据需要调整文件扩展名
                file_path = os.path.join(root, file)
                try:
                    workbook = load_workbook(file_path, data_only=True)  # data_only=True用于读取公式的结果而不是公式本身
                    sheet = workbook.active  # 获取活动工作表，或者可以使用 workbook['SheetName'] 指定工作表
                    row = 2  # 从第二行开始
                    col_start = 3  # 从第三列开始
                    while True:
                        cell_value = sheet.cell(row=row, column=col_start).value
                        if cell_value is None:
                            break  # 遇到空单元格，停止循环
                        values.append(cell_value)
                        col_start += 1  # 移动到下一列
                    # 如果需要，可以在这里处理每个文件的值，或者将它们存储在一个字典中以文件名为键
                except Exception as e:
                    print(f"无法读取文件 {file_path}。错误: {e}")
    return values

# 使用示例

upload_dir = os.path.join(os.getcwd(), 'projects')
directory_path = upload_dir  # 替换为你的Excel文件所在的目录路径
all_values = get_excel_values(directory_path)
print(all_values)